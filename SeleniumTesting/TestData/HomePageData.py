import openpyxl


class HomePageData:
    
    test_HomePage_data = [{'name': 'Kev', 'email': 'test@gmail.com', 'gender': 'Male'},
                          {'name': 'Kelvin', 'email': 'test123124@gmail.com', 'gender': 'Male'}
                         ]
    
    # Creating a static method you do not -
    # require to add self to method/function
    @staticmethod
    def getTestData(testName):

        Dict = {}

        book = openpyxl.load_workbook('E:\Study\Python Selenium\Selenium\Excel\PythonDemo.xlsx')

        sheet = book.active


        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == testName:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
